"""
定时报表与数据分析模块
每周三凌晨自动统计并生成运维分析报告
"""
import os
import json
import uuid
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime, timedelta
from collections import defaultdict

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
    Image, PageBreak, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .database import get_db
from .logger import get_logger
from .config import get_config
from .constants import DeploymentStatus, RiskLevel, OperationType
from .audit import get_audit_logger

logger = get_logger(__name__)


class WeeklyReportGenerator:
    """周度运维报表生成器"""
    
    def __init__(self):
        self.config = get_config()
        self.db = get_db()
        self.output_dir = self.config.get('reporting.output_dir', './reports')
        os.makedirs(self.output_dir, exist_ok=True)
        
        chart_config = self.config.get('reporting.charts', {})
        self.chart_dpi = chart_config.get('dpi', 300)
        self.figure_size = tuple(chart_config.get('figure_size', [12, 8]))
        plt.style.use(chart_config.get('style', 'ggplot'))
        
    def _get_report_period(self) -> Tuple[str, datetime, datetime]:
        """获取报表统计周期（上周三到本周二）"""
        today = datetime.now()
        end_date = today - timedelta(days=(today.weekday() + 1) % 7)
        start_date = end_date - timedelta(days=6)
        
        period_str = f"{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}"
        return period_str, start_date, end_date
    
    def _calculate_stats(self, start_date: datetime, 
                         end_date: datetime) -> Dict[str, Any]:
        """计算统计数据"""
        start_str = start_date.strftime('%Y-%m-%d 00:00:00')
        end_str = end_date.strftime('%Y-%m-%d 23:59:59')
        
        releases = self.db.query('''
            SELECT * FROM release_requests 
            WHERE created_at >= ? AND created_at <= ?
            ORDER BY created_at
        ''', (start_str, end_str))
        
        total_releases = len(releases)
        successful_releases = sum(
            1 for r in releases 
            if r['status'] in [
                DeploymentStatus.FULL_DEPLOYED.value,
                DeploymentStatus.GRAY_OBSERVING.value
            ]
        )
        
        publish_success_rate = (successful_releases / total_releases * 100) if total_releases > 0 else 100
        
        rollbacks = self.db.query('''
            SELECT * FROM rollback_records 
            WHERE rollback_time >= ? AND rollback_time <= ?
            ORDER BY rollback_time
        ''', (start_str, end_str))
        
        emergency_rollback_count = len(rollbacks)
        
        approvals = self.db.query('''
            SELECT 
                ar.request_id,
                MIN(ar.approved_at) as first_approval,
                MAX(ar.approved_at) as last_approval
            FROM approval_records ar
            INNER JOIN release_requests rr ON ar.request_id = rr.request_id
            WHERE ar.approval_status = 'APPROVED'
              AND ar.approved_at >= ? AND ar.approved_at <= ?
            GROUP BY ar.request_id
        ''', (start_str, end_str))
        
        approval_durations = []
        for app in approvals:
            if app['first_approval'] and app['last_approval']:
                first = datetime.strptime(app['first_approval'], '%Y-%m-%d %H:%M:%S')
                last = datetime.strptime(app['last_approval'], '%Y-%m-%d %H:%M:%S')
                duration_minutes = (last - first).total_seconds() / 60
                approval_durations.append(duration_minutes)
        
        avg_approval_duration = (
            sum(approval_durations) / len(approval_durations) 
            if approval_durations else 0
        )
        
        risk_distribution = defaultdict(int)
        for r in releases:
            risk_distribution[r['risk_level']] += 1
        
        daily_releases = defaultdict(int)
        for r in releases:
            day = r['created_at'][:10]
            daily_releases[day] += 1
        
        deployment_stages = defaultdict(int)
        deploy_records = self.db.query('''
            SELECT stage_name, COUNT(*) as count 
            FROM deployment_records 
            WHERE start_time >= ? AND start_time <= ?
            GROUP BY stage_name
        ''', (start_str, end_str))
        
        for d in deploy_records:
            deployment_stages[d['stage_name']] = d['count']
        
        return {
            'period': {
                'start': start_str,
                'end': end_str
            },
            'total_releases': total_releases,
            'successful_releases': successful_releases,
            'failed_releases': total_releases - successful_releases,
            'publish_success_rate': round(publish_success_rate, 2),
            'emergency_rollback_count': emergency_rollback_count,
            'avg_approval_duration_minutes': round(avg_approval_duration, 2),
            'risk_distribution': dict(risk_distribution),
            'daily_releases': dict(daily_releases),
            'deployment_stages': dict(deployment_stages),
            'releases': releases,
            'rollbacks': rollbacks
        }
    
    def _generate_charts(self, stats: Dict[str, Any], 
                        report_id: str) -> Dict[str, str]:
        """生成统计图表"""
        chart_paths = {}
        chart_dir = os.path.join(self.output_dir, 'charts')
        os.makedirs(chart_dir, exist_ok=True)
        
        risk_data = stats['risk_distribution']
        if risk_data:
            fig, ax = plt.subplots(figsize=self.figure_size)
            labels = [RiskLevel(k).value for k in risk_data.keys()]
            values = list(risk_data.values())
            colors_pie = ['#FF6B6B', '#4ECDC4']
            
            wedges, texts, autotexts = ax.pie(
                values, labels=labels, autopct='%1.1f%%',
                colors=colors_pie, startangle=90
            )
            ax.set_title('版本发布风险等级分布', fontsize=16, fontweight='bold')
            plt.tight_layout()
            
            path = os.path.join(chart_dir, f'{report_id}_risk_distribution.png')
            plt.savefig(path, dpi=self.chart_dpi, bbox_inches='tight')
            plt.close()
            chart_paths['risk_distribution'] = path
        
        daily_data = stats['daily_releases']
        if daily_data:
            fig, ax = plt.subplots(figsize=self.figure_size)
            days = sorted(daily_data.keys())
            counts = [daily_data[d] for d in days]
            
            ax.bar(days, counts, color='#3498DB', alpha=0.8)
            ax.set_title('每日发布次数趋势', fontsize=16, fontweight='bold')
            ax.set_xlabel('日期', fontsize=12)
            ax.set_ylabel('发布次数', fontsize=12)
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            path = os.path.join(chart_dir, f'{report_id}_daily_trend.png')
            plt.savefig(path, dpi=self.chart_dpi, bbox_inches='tight')
            plt.close()
            chart_paths['daily_trend'] = path
        
        if 'successful_releases' in stats:
            fig, ax = plt.subplots(figsize=self.figure_size)
            
            success_data = [
                ('成功发布', stats['successful_releases']),
                ('失败/回滚', stats['failed_releases'] + stats['emergency_rollback_count'])
            ]
            labels, values = zip(*success_data)
            colors_bar = ['#27AE60', '#E74C3C']
            
            bars = ax.bar(labels, values, color=colors_bar, alpha=0.8)
            ax.set_title('发布成功率统计', fontsize=16, fontweight='bold')
            ax.set_ylabel('次数', fontsize=12)
            
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                       f'{int(height)}', ha='center', va='bottom', fontsize=12)
            
            plt.tight_layout()
            
            path = os.path.join(chart_dir, f'{report_id}_success_rate.png')
            plt.savefig(path, dpi=self.chart_dpi, bbox_inches='tight')
            plt.close()
            chart_paths['success_rate'] = path
        
        metrics_data = {
            '发布成功率(%)': stats['publish_success_rate'],
            '紧急回滚次数': stats['emergency_rollback_count'],
            '平均审批时长(分钟)': stats['avg_approval_duration_minutes']
        }
        
        if any(metrics_data.values()):
            fig, ax = plt.subplots(figsize=self.figure_size)
            
            metrics = list(metrics_data.keys())
            values = list(metrics_data.values())
            
            ax.barh(metrics, values, color=['#3498DB', '#E74C3C', '#F39C12'], alpha=0.8)
            ax.set_title('核心运维指标', fontsize=16, fontweight='bold')
            ax.set_xlabel('数值', fontsize=12)
            
            for i, v in enumerate(values):
                ax.text(v + 0.1, i, str(v), va='center', fontsize=12)
            
            plt.tight_layout()
            
            path = os.path.join(chart_dir, f'{report_id}_metrics.png')
            plt.savefig(path, dpi=self.chart_dpi, bbox_inches='tight')
            plt.close()
            chart_paths['metrics'] = path
        
        return chart_paths
    
    def _generate_pdf(self, stats: Dict[str, Any], 
                      chart_paths: Dict[str, str],
                      report_id: str, period_str: str) -> str:
        """生成PDF报告"""
        pdf_path = os.path.join(self.output_dir, f'{report_id}_week_report.pdf')
        
        doc = SimpleDocTemplate(
            pdf_path, pagesize=A4,
            rightMargin=0.5*inch, leftMargin=0.5*inch,
            topMargin=0.5*inch, bottomMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=18, textColor=colors.HexColor('#2C3E50'),
            alignment=TA_CENTER, spaceAfter=20
        )
        section_style = ParagraphStyle(
            'SectionTitle', parent=styles['Heading2'],
            fontSize=14, textColor=colors.HexColor('#3498DB'),
            spaceBefore=15, spaceAfter=10
        )
        normal_style = styles['Normal']
        
        story = []
        
        story.append(Paragraph('MES系统运维周度分析报告', title_style))
        story.append(Paragraph(
            f'统计周期: {stats["period"]["start"]} 至 {stats["period"]["end"]}',
            ParagraphStyle('CenterText', parent=normal_style, alignment=TA_CENTER, fontSize=12)
        ))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#3498DB')))
        story.append(Spacer(1, 0.2*inch))
        
        story.append(Paragraph('一、核心指标概览', section_style))
        
        summary_data = [
            ['指标', '数值', '说明'],
            ['总发布次数', str(stats['total_releases']), '本周申请发布总数'],
            ['成功发布次数', str(stats['successful_releases']), '完成全量部署的版本数'],
            ['发布成功率', f'{stats["publish_success_rate"]}%', '成功发布/总发布'],
            ['紧急回滚次数', str(stats['emergency_rollback_count']), '监控触发的自动回滚'],
            ['平均审批时长', f'{stats["avg_approval_duration_minutes"]}分钟', '从创建到完成审批平均耗时']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECF0F1')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9F9')])
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        if 'success_rate' in chart_paths:
            story.append(Paragraph('二、发布成功率分析', section_style))
            story.append(Image(chart_paths['success_rate'], width=6*inch, height=4*inch))
            story.append(Spacer(1, 0.2*inch))
        
        if 'risk_distribution' in chart_paths:
            story.append(Paragraph('三、风险等级分布', section_style))
            story.append(Image(chart_paths['risk_distribution'], width=6*inch, height=4*inch))
            story.append(Spacer(1, 0.2*inch))
        
        if 'daily_trend' in chart_paths:
            story.append(Paragraph('四、每日发布趋势', section_style))
            story.append(Image(chart_paths['daily_trend'], width=6*inch, height=4*inch))
            story.append(Spacer(1, 0.2*inch))
        
        if 'metrics' in chart_paths:
            story.append(Paragraph('五、运维指标详情', section_style))
            story.append(Image(chart_paths['metrics'], width=6*inch, height=4*inch))
            story.append(PageBreak())
        
        story.append(Paragraph('六、版本发布明细', section_style))
        
        if stats['releases']:
            release_data = [['申请ID', '版本号', '风险等级', '状态', '申请人', '创建时间']]
            for r in stats['releases']:
                release_data.append([
                    r['request_id'],
                    r['version'],
                    r['risk_level'],
                    r['status'],
                    r['applicant'],
                    r['created_at']
                ])
            
            release_table = Table(release_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1.2*inch, 1*inch, 1.3*inch])
            release_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9F9')])
            ]))
            story.append(release_table)
        else:
            story.append(Paragraph('本周无版本发布记录', normal_style))
        
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph('七、趋势分析与建议', section_style))
        
        suggestions = self._generate_suggestions(stats)
        for i, suggestion in enumerate(suggestions, 1):
            story.append(Paragraph(f'{i}. {suggestion}', normal_style))
            story.append(Spacer(1, 0.1*inch))
        
        story.append(Spacer(1, 0.5*inch))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#BDC3C7')))
        story.append(Paragraph(
            f'报告生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | 报告编号: {report_id}',
            ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.gray)
        ))
        
        doc.build(story)
        logger.info(f"PDF报告已生成: {pdf_path}")
        
        return pdf_path
    
    def _generate_excel(self, stats: Dict[str, Any], 
                        report_id: str) -> str:
        """生成Excel版本上线历史明细报表"""
        excel_path = os.path.join(self.output_dir, f'{report_id}_release_history.xlsx')
        
        wb = Workbook()
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws1 = wb.active
        ws1.title = "发布概览"
        
        ws1['A1'] = "MES系统版本上线历史明细报表"
        ws1.merge_cells('A1:F1')
        ws1['A1'].font = Font(bold=True, size=16, color="2C3E50")
        ws1['A1'].alignment = center_align
        
        ws1['A3'] = f"统计周期: {stats['period']['start']} 至 {stats['period']['end']}"
        ws1.merge_cells('A3:F3')
        
        summary_headers = ['指标', '数值', '说明']
        for col, header in enumerate(summary_headers, 1):
            cell = ws1.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        summary_data = [
            ['总发布次数', stats['total_releases'], '本周申请发布总数'],
            ['成功发布次数', stats['successful_releases'], '完成全量部署的版本数'],
            ['发布成功率', f"{stats['publish_success_rate']}%", '成功发布/总发布'],
            ['紧急回滚次数', stats['emergency_rollback_count'], '监控触发的自动回滚'],
            ['平均审批时长', f"{stats['avg_approval_duration_minutes']}分钟", '从创建到完成审批平均耗时']
        ]
        
        for row_idx, row_data in enumerate(summary_data, 6):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = left_align if col_idx == 3 else center_align
                cell.border = thin_border
        
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 40
        
        ws2 = wb.create_sheet("发布明细")
        
        release_headers = [
            '申请ID', '版本号', '风险等级', '状态', '申请人',
            '部门', '描述', '创建时间', '更新时间'
        ]
        
        for col, header in enumerate(release_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for row_idx, release in enumerate(stats['releases'], 2):
            ws2.cell(row=row_idx, column=1, value=release['request_id']).border = thin_border
            ws2.cell(row=row_idx, column=2, value=release['version']).border = thin_border
            ws2.cell(row=row_idx, column=3, value=release['risk_level']).border = thin_border
            ws2.cell(row=row_idx, column=4, value=release['status']).border = thin_border
            ws2.cell(row=row_idx, column=5, value=release['applicant']).border = thin_border
            ws2.cell(row=row_idx, column=6, value=release.get('department', '')).border = thin_border
            ws2.cell(row=row_idx, column=7, value=release.get('description', '')).border = thin_border
            ws2.cell(row=row_idx, column=8, value=release['created_at']).border = thin_border
            ws2.cell(row=row_idx, column=9, value=release['updated_at']).border = thin_border
        
        for col in range(1, 10):
            ws2.column_dimensions[get_column_letter(col)].width = 20
        
        ws3 = wb.create_sheet("回滚记录")
        
        rollback_headers = [
            '回滚ID', '申请ID', '回滚原因', '从版本', '到版本',
            '受影响产线', '预估不良品', '回滚时间'
        ]
        
        for col, header in enumerate(rollback_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for row_idx, rollback in enumerate(stats['rollbacks'], 2):
            affected_lines = json.loads(rollback['affected_lines']) if rollback['affected_lines'] else []
            ws3.cell(row=row_idx, column=1, value=rollback['id']).border = thin_border
            ws3.cell(row=row_idx, column=2, value=rollback['request_id']).border = thin_border
            ws3.cell(row=row_idx, column=3, value=rollback['rollback_reason']).border = thin_border
            ws3.cell(row=row_idx, column=4, value=rollback['from_version']).border = thin_border
            ws3.cell(row=row_idx, column=5, value=rollback['to_version']).border = thin_border
            ws3.cell(row=row_idx, column=6, value=", ".join(affected_lines)).border = thin_border
            ws3.cell(row=row_idx, column=7, value=rollback['estimated_defect_count']).border = thin_border
            ws3.cell(row=row_idx, column=8, value=rollback['rollback_time']).border = thin_border
        
        for col in range(1, 9):
            ws3.column_dimensions[get_column_letter(col)].width = 20
        
        wb.save(excel_path)
        logger.info(f"Excel报表已生成: {excel_path}")
        
        return excel_path
    
    def _generate_suggestions(self, stats: Dict[str, Any]) -> List[str]:
        """生成运维建议"""
        suggestions = []
        
        if stats['publish_success_rate'] < 95:
            suggestions.append(
                f"本周发布成功率为{stats['publish_success_rate']}%，低于95%的目标值。"
                "建议加强前置校验环节，增加代码评审和测试覆盖度。"
            )
        
        if stats['emergency_rollback_count'] > 2:
            suggestions.append(
                f"本周发生{stats['emergency_rollback_count']}次紧急回滚。"
                "建议延长灰度观察周期，优化监控阈值设置，提升问题发现的及时性。"
            )
        
        if stats['avg_approval_duration_minutes'] > 120:
            suggestions.append(
                f"本周平均审批时长为{stats['avg_approval_duration_minutes']}分钟，超过2小时。"
                "建议优化审批流程，考虑增加审批超时自动升级机制。"
            )
        
        if stats.get('risk_distribution', {}).get('L2_URGENT', 0) > 3:
            suggestions.append(
                "本周紧急版本发布较多，建议加强版本规划，"
                "减少临时紧急发布，降低生产环境风险。"
            )
        
        if not suggestions:
            suggestions.append(
                "本周各项运维指标表现良好，发布成功率达标，"
                "建议继续保持当前的发布节奏和质量管控措施。"
            )
        
        suggestions.append(
            "建议下周三前完成本周故障的根因分析和整改措施落实，"
            "形成闭环管理，避免同类问题重复发生。"
        )
        
        return suggestions
    
    def generate_weekly_report(self, start_date: datetime = None, 
                                end_date: datetime = None) -> Dict[str, Any]:
        """
        生成完整的周度报告
        
        Args:
            start_date: 开始日期（可选，默认自动计算上周）
            end_date: 结束日期（可选，默认自动计算本周二）
            
        Returns:
            完整的报告结果字典
        """
        report_id = f"WEEKLY-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        if start_date is not None and end_date is not None:
            period_str = f"{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}"
        else:
            period_str, start_date, end_date = self._get_report_period()
        
        logger.info(f"开始生成周度报告: {report_id}, 周期: {period_str}")
        
        stats = self._calculate_stats(start_date, end_date)
        chart_paths = self._generate_charts(stats, report_id)
        pdf_path = self._generate_pdf(stats, chart_paths, report_id, period_str)
        excel_path = self._generate_excel(stats, report_id)
        
        self.db.execute('''
            INSERT INTO weekly_reports 
            (report_id, report_period, publish_success_rate, 
             emergency_rollback_count, avg_approval_duration, 
             pdf_path, excel_path, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            report_id, period_str, stats['publish_success_rate'],
            stats['emergency_rollback_count'], stats['avg_approval_duration_minutes'],
            pdf_path, excel_path, datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ))
        
        get_audit_logger().log(
            operation_type=OperationType.SYSTEM_CONFIG,
            operator="system",
            request_params={
                "report_id": report_id,
                "period": period_str
            },
            response_result={
                "publish_success_rate": stats['publish_success_rate'],
                "emergency_rollback_count": stats['emergency_rollback_count']
            },
            status="SUCCESS"
        )
        
        report = {
            'report_id': report_id,
            'report_period': period_str,
            'stats': stats,
            'chart_paths': chart_paths,
            'pdf_path': pdf_path,
            'excel_path': excel_path,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        report_path = os.path.join(self.output_dir, f'{report_id}.json')
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        logger.info(f"周度报告生成完成: {report_id}")
        
        # 返回元组以兼容解包，同时包含完整信息
        return (report_id, pdf_path, excel_path, report)


def get_report_generator() -> WeeklyReportGenerator:
    """获取报表生成器"""
    return WeeklyReportGenerator()
