"""
历史记录查询与导出模块
支持多条件组合检索和批量导出
"""
import os
import json
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .database import get_db
from .logger import get_logger
from .config import get_config
from .constants import OperationType
from .audit import get_audit_logger, audit_operation

logger = get_logger(__name__)


class QueryExportManager:
    """查询与导出管理器"""
    
    def __init__(self):
        self.config = get_config()
        self.db = get_db()
        self.export_dir = self.config.get('reporting.output_dir', './exports')
        os.makedirs(self.export_dir, exist_ok=True)
    
    def query_release_records(self, 
                              start_time = None,
                              end_time = None,
                              start_date = None,
                              end_date = None,
                              workshop: str = None,
                              version: str = None,
                              status: str = None,
                              risk_level: str = None,
                              applicant: str = None,
                              limit: int = 500) -> List[Dict[str, Any]]:
        """
        多条件组合查询发布记录
        
        Args:
            start_time: 上线发布开始时间（或start_date）
            end_time: 上线发布结束时间（或end_date）
            start_date: 上线发布开始日期（兼容参数）
            end_date: 上线发布结束日期（兼容参数）
            workshop: 生产车间模块
            version: MES系统版本号
            status: 发布状态
            risk_level: 风险等级
            applicant: 申请人
            limit: 返回条数限制
            
        Returns:
            发布记录列表
        """
        if start_date and not start_time:
            start_time = start_date
        if end_date and not end_time:
            end_time = end_date
        
        if isinstance(start_time, datetime):
            start_time = start_time.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(end_time, datetime):
            end_time = end_time.strftime('%Y-%m-%d %H:%M:%S')
        
        sql = '''
            SELECT DISTINCT rr.*, 
                   pr.check_result as pre_check_status,
                   GROUP_CONCAT(ar.approver_name || ':' || ar.approval_status, ';') as approval_progress
            FROM release_requests rr
            LEFT JOIN pre_check_records pr ON rr.request_id = pr.request_id
            LEFT JOIN approval_records ar ON rr.request_id = ar.request_id
            WHERE 1=1
        '''
        params = []
        
        if start_time:
            sql += " AND rr.created_at >= ?"
            params.append(start_time)
        if end_time:
            sql += " AND rr.created_at <= ?"
            params.append(end_time)
        if version:
            sql += " AND rr.version LIKE ?"
            params.append(f"%{version}%")
        if status:
            sql += " AND rr.status = ?"
            params.append(status)
        if risk_level:
            sql += " AND rr.risk_level = ?"
            params.append(risk_level)
        if applicant:
            sql += " AND rr.applicant LIKE ?"
            params.append(f"%{applicant}%")
        if workshop:
            sql += " AND rr.target_production_lines LIKE ?"
            params.append(f"%{workshop}%")
        
        sql += " GROUP BY rr.id ORDER BY rr.id DESC LIMIT ?"
        params.append(limit)
        
        results = self.db.query(sql, tuple(params))
        
        for result in results:
            if result.get('target_production_lines'):
                try:
                    result['target_production_lines'] = json.loads(result['target_production_lines'])
                except:
                    pass
            if result.get('pre_check_result'):
                try:
                    result['pre_check_result'] = json.loads(result['pre_check_result'])
                except:
                    pass
        
        logger.info(f"查询到 {len(results)} 条发布记录")
        return results
    
    def query_rollback_records(self,
                               start_time: str = None,
                               end_time: str = None,
                               version: str = None,
                               workshop: str = None,
                               limit: int = 200) -> List[Dict[str, Any]]:
        """查询回滚记录"""
        sql = '''
            SELECT rb.*, rr.version, rr.applicant
            FROM rollback_records rb
            LEFT JOIN release_requests rr ON rb.request_id = rr.request_id
            WHERE 1=1
        '''
        params = []
        
        if start_time:
            sql += " AND rb.rollback_time >= ?"
            params.append(start_time)
        if end_time:
            sql += " AND rb.rollback_time <= ?"
            params.append(end_time)
        if version:
            sql += " AND (rb.from_version LIKE ? OR rb.to_version LIKE ?)"
            params.extend([f"%{version}%", f"%{version}%"])
        if workshop:
            sql += " AND rb.affected_lines LIKE ?"
            params.append(f"%{workshop}%")
        
        sql += " ORDER BY rb.id DESC LIMIT ?"
        params.append(limit)
        
        results = self.db.query(sql, tuple(params))
        
        for result in results:
            if result.get('affected_lines'):
                try:
                    result['affected_lines'] = json.loads(result['affected_lines'])
                except:
                    pass
            if result.get('trigger_metrics'):
                try:
                    result['trigger_metrics'] = json.loads(result['trigger_metrics'])
                except:
                    pass
            if result.get('root_cause'):
                try:
                    result['root_cause'] = json.loads(result['root_cause'])
                except:
                    pass
        
        logger.info(f"查询到 {len(results)} 条回滚记录")
        return results
    
    def query_approval_records(self,
                               request_id: str = None,
                               approver_role: str = None,
                               approval_status: str = None,
                               start_time: str = None,
                               end_time: str = None,
                               limit: int = 500) -> List[Dict[str, Any]]:
        """查询审批记录"""
        sql = '''
            SELECT ar.*, rr.version, rr.risk_level, rr.applicant, rr.description
            FROM approval_records ar
            LEFT JOIN release_requests rr ON ar.request_id = rr.request_id
            WHERE 1=1
        '''
        params = []
        
        if request_id:
            sql += " AND ar.request_id = ?"
            params.append(request_id)
        if approver_role:
            sql += " AND ar.approver_role = ?"
            params.append(approver_role)
        if approval_status:
            sql += " AND ar.approval_status = ?"
            params.append(approval_status)
        if start_time:
            sql += " AND ar.approved_at >= ?"
            params.append(start_time)
        if end_time:
            sql += " AND ar.approved_at <= ?"
            params.append(end_time)
        
        sql += " ORDER BY ar.id DESC LIMIT ?"
        params.append(limit)
        
        results = self.db.query(sql, tuple(params))
        logger.info(f"查询到 {len(results)} 条审批记录")
        return results
    
    def query_production_line_status(self,
                                      workshop: str = None,
                                      line_name: str = None,
                                      auto_enabled: bool = None) -> List[Dict[str, Any]]:
        """查询产线状态"""
        sql = "SELECT * FROM production_line_status WHERE 1=1"
        params = []
        
        if workshop:
            sql += " AND line_name LIKE ?"
            params.append(f"%{workshop}%")
        if line_name:
            sql += " AND line_name = ?"
            params.append(line_name)
        if auto_enabled is not None:
            sql += " AND auto_production_enabled = ?"
            params.append(1 if auto_enabled else 0)
        
        sql += " ORDER BY line_name"
        
        results = self.db.query(sql, tuple(params))
        logger.info(f"查询到 {len(results)} 条产线状态")
        return results
    
    def query_audit_logs(self,
                         operation_type: str = None,
                         operator: str = None,
                         start_time: str = None,
                         end_time: str = None,
                         status: str = None,
                         limit: int = 500) -> List[Dict[str, Any]]:
        """查询审计日志"""
        sql = "SELECT * FROM audit_logs WHERE 1=1"
        params = []
        
        if operation_type:
            sql += " AND operation_type = ?"
            params.append(operation_type)
        if operator:
            sql += " AND operator LIKE ?"
            params.append(f"%{operator}%")
        if start_time:
            sql += " AND created_at >= ?"
            params.append(start_time)
        if end_time:
            sql += " AND created_at <= ?"
            params.append(end_time)
        if status:
            sql += " AND status = ?"
            params.append(status)
        
        sql += " ORDER BY id DESC LIMIT ?"
        params.append(limit)
        
        results = self.db.query(sql, tuple(params))
        
        for result in results:
            if result.get('request_params'):
                try:
                    result['request_params'] = json.loads(result['request_params'])
                except:
                    pass
            if result.get('response_result'):
                try:
                    result['response_result'] = json.loads(result['response_result'])
                except:
                    pass
        
        logger.info(f"查询到 {len(results)} 条审计日志")
        return results
    
    def query_drill_records(self,
                            drill_type: str = None,
                            status: str = None,
                            operator: str = None,
                            start_time: str = None,
                            end_time: str = None,
                            limit: int = 100) -> List[Dict[str, Any]]:
        """查询应急演练记录"""
        sql = "SELECT * FROM emergency_drills WHERE 1=1"
        params = []
        
        if drill_type:
            sql += " AND drill_type = ?"
            params.append(drill_type)
        if status:
            sql += " AND status = ?"
            params.append(status)
        if operator:
            sql += " AND operator LIKE ?"
            params.append(f"%{operator}%")
        if start_time:
            sql += " AND started_at >= ?"
            params.append(start_time)
        if end_time:
            sql += " AND started_at <= ?"
            params.append(end_time)
        
        sql += " ORDER BY id DESC LIMIT ?"
        params.append(limit)
        
        results = self.db.query(sql, tuple(params))
        
        for result in results:
            if result.get('trigger_scenario'):
                try:
                    result['trigger_scenario'] = json.loads(result['trigger_scenario'])
                except:
                    pass
            if result.get('drill_result'):
                try:
                    result['drill_result'] = json.loads(result['drill_result'])
                except:
                    pass
            if result.get('improvements'):
                try:
                    result['improvements'] = json.loads(result['improvements'])
                except:
                    pass
        
        logger.info(f"查询到 {len(results)} 条演练记录")
        return results
    
    @audit_operation(OperationType.SYSTEM_CONFIG, lambda *args, **kwargs: kwargs.get('operator', args[1] if len(args) > 1 else 'system'))
    def export_to_excel(self, operator: str, data_type: str,
                        records: List[Dict[str, Any]],
                        filename: str = None) -> str:
        """
        批量导出数据到Excel
        
        Args:
            operator: 操作人
            data_type: 数据类型 (releases/approvals/rollbacks/audits/drills)
            records: 要导出的数据记录
            filename: 文件名（可选）
            
        Returns:
            导出文件路径
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"{data_type}_{timestamp}.xlsx"
        
        export_path = os.path.join(self.export_dir, filename)
        
        wb = Workbook()
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        sheet_configs = {
            'releases': {
                'title': '发布记录',
                'headers': ['申请ID', '版本号', '风险等级', '状态', '申请人', '部门', '描述', '创建时间', '更新时间']
            },
            'approvals': {
                'title': '审批单据',
                'headers': ['申请ID', '版本号', '审批人角色', '审批人', '审批状态', '审批意见', '审批时间']
            },
            'rollbacks': {
                'title': '故障处置记录',
                'headers': ['回滚ID', '申请ID', '回滚原因', '从版本', '到版本', '受影响产线', '预估不良品', '回滚时间']
            },
            'audits': {
                'title': '审计日志',
                'headers': ['审计ID', '操作类型', '操作人', 'IP地址', '状态', '耗时(ms)', '操作时间']
            },
            'drills': {
                'title': '应急演练记录',
                'headers': ['演练ID', '演练名称', '演练类型', '状态', '操作人', '开始时间', '结束时间']
            },
            'production_lines': {
                'title': '产线停机记录',
                'headers': ['产线名称', '是否运行', '自动生产', '当前版本', '兜底模式', '最后心跳']
            }
        }
        
        config = sheet_configs.get(data_type, sheet_configs['releases'])
        
        ws = wb.active
        ws.title = config['title']
        
        for col, header in enumerate(config['headers'], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        row_mappers = {
            'releases': lambda r: [
                r.get('request_id', ''),
                r.get('version', ''),
                r.get('risk_level', ''),
                r.get('status', ''),
                r.get('applicant', ''),
                r.get('department', ''),
                r.get('description', ''),
                r.get('created_at', ''),
                r.get('updated_at', '')
            ],
            'approvals': lambda r: [
                r.get('request_id', ''),
                r.get('version', ''),
                r.get('approver_role', ''),
                r.get('approver_name', ''),
                r.get('approval_status', ''),
                r.get('approval_comment', ''),
                r.get('approved_at', '')
            ],
            'rollbacks': lambda r: [
                r.get('id', ''),
                r.get('request_id', ''),
                r.get('rollback_reason', ''),
                r.get('from_version', ''),
                r.get('to_version', ''),
                ', '.join(r.get('affected_lines', []) if isinstance(r.get('affected_lines'), list) else []),
                r.get('estimated_defect_count', ''),
                r.get('rollback_time', '')
            ],
            'audits': lambda r: [
                r.get('audit_id', ''),
                r.get('operation_type', ''),
                r.get('operator', ''),
                r.get('ip_address', ''),
                r.get('status', ''),
                r.get('duration_ms', ''),
                r.get('created_at', '')
            ],
            'drills': lambda r: [
                r.get('drill_id', ''),
                r.get('drill_name', ''),
                r.get('drill_type', ''),
                r.get('status', ''),
                r.get('operator', ''),
                r.get('started_at', ''),
                r.get('completed_at', '')
            ],
            'production_lines': lambda r: [
                r.get('line_name', ''),
                '是' if r.get('is_running') else '否',
                '启用' if r.get('auto_production_enabled') else '禁用',
                r.get('current_version', ''),
                r.get('fallback_mode', ''),
                r.get('last_heartbeat', '')
            ]
        }
        
        mapper = row_mappers.get(data_type, row_mappers['releases'])
        
        for row_idx, record in enumerate(records, 2):
            row_data = mapper(record)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx > 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for col in range(1, len(config['headers']) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        wb.save(export_path)
        
        get_audit_logger().log(
            operation_type=OperationType.SYSTEM_CONFIG,
            operator=operator,
            request_params={
                "data_type": data_type,
                "record_count": len(records),
                "filename": filename
            },
            response_result={"export_path": export_path},
            status="SUCCESS"
        )
        
        logger.info(f"已导出 {len(records)} 条 {data_type} 记录到 {export_path}")
        
        return export_path
    
    def export_records(self, export_type: str, output_path: str,
                        start_date: datetime = None, end_date: datetime = None,
                        operator: str = 'system') -> str:
        """
        便捷方法：导出记录到Excel（兼容main.py调用）
        
        Args:
            export_type: 导出类型 (all/releases/approvals/rollbacks/audits/drills/production_lines)
            output_path: 输出文件路径
            start_date: 开始日期
            end_date: 结束日期
            operator: 操作人
            
        Returns:
            导出文件路径
        """
        import os
        output_dir = os.path.dirname(output_path) if os.path.dirname(output_path) else '.'
        os.makedirs(output_dir, exist_ok=True)
        
        if start_date:
            start_time = start_date.strftime('%Y-%m-%d 00:00:00')
        else:
            start_time = None
            
        if end_date:
            end_time = end_date.strftime('%Y-%m-%d 23:59:59')
        else:
            end_time = None
        
        if export_type == 'all':
            result = self.batch_export(
                operator=operator,
                start_time=start_time,
                end_time=end_time
            )
            files = list(result.get('export_files', {}).values())
            return files[0] if files else output_path
        else:
            query_funcs = {
                'releases': self.query_release_records,
                'approvals': self.query_approval_records,
                'rollbacks': self.query_rollback_records,
                'drills': self.query_drill_records,
                'production_lines': self.query_production_line_status,
                'audits': self.query_audit_logs
            }
            
            query_func = query_funcs.get(export_type)
            if query_func:
                records = query_func(start_time=start_time, end_time=end_time)
            else:
                records = []
            
            return self._do_export(
                operator=operator,
                data_type=export_type,
                records=records,
                output_path=output_path
            )
    
    def _do_export(self, operator: str, data_type: str,
                      records: List[Dict[str, Any]],
                      output_path: str) -> str:
        """实际导出Excel"""
        return self.export_to_excel(
            operator=operator,
            data_type=data_type,
            records=records,
            filename=os.path.basename(output_path)
        )
    
    def batch_export(self, operator: str,
                     start_time: str = None,
                     end_time: str = None,
                     workshop: str = None,
                     version: str = None) -> Dict[str, Any]:
        """
        批量导出审批单据、故障处置、产线停机记录
        
        Args:
            operator: 操作人
            start_time: 开始时间
            end_time: 结束时间
            workshop: 车间
            version: 版本号
            
        Returns:
            导出结果字典
        """
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        
        releases = self.query_release_records(
            start_time=start_time, end_time=end_time,
            workshop=workshop, version=version
        )
        
        approvals = self.query_approval_records(
            start_time=start_time, end_time=end_time
        )
        
        rollbacks = self.query_rollback_records(
            start_time=start_time, end_time=end_time,
            version=version, workshop=workshop
        )
        
        production_lines = self.query_production_line_status(workshop=workshop)
        
        exports = {}
        
        if releases:
            exports['releases'] = self.export_to_excel(
                operator, 'releases', releases,
                f"release_records_{timestamp}.xlsx"
            )
        
        if approvals:
            exports['approvals'] = self.export_to_excel(
                operator, 'approvals', approvals,
                f"approval_records_{timestamp}.xlsx"
            )
        
        if rollbacks:
            exports['rollbacks'] = self.export_to_excel(
                operator, 'rollbacks', rollbacks,
                f"rollback_records_{timestamp}.xlsx"
            )
        
        if production_lines:
            exports['production_lines'] = self.export_to_excel(
                operator, 'production_lines', production_lines,
                f"production_line_status_{timestamp}.xlsx"
            )
        
        return {
            'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'operator': operator,
            'criteria': {
                'start_time': start_time,
                'end_time': end_time,
                'workshop': workshop,
                'version': version
            },
            'record_counts': {
                'releases': len(releases),
                'approvals': len(approvals),
                'rollbacks': len(rollbacks),
                'production_lines': len(production_lines)
            },
            'export_files': exports
        }
    
    def get_statistics_summary(self,
                               start_time: str = None,
                               end_time: str = None) -> Dict[str, Any]:
        """获取统计概要"""
        if not start_time:
            start_time = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d 00:00:00')
        if not end_time:
            end_time = datetime.now().strftime('%Y-%m-%d 23:59:59')
        
        releases = self.query_release_records(start_time, end_time, limit=10000)
        rollbacks = self.query_rollback_records(start_time, end_time, limit=10000)
        
        status_stats = {}
        for r in releases:
            status = r['status']
            status_stats[status] = status_stats.get(status, 0) + 1
        
        risk_stats = {}
        for r in releases:
            risk = r['risk_level']
            risk_stats[risk] = risk_stats.get(risk, 0) + 1
        
        success_count = sum(1 for r in releases if r['status'] == 'FULL_DEPLOYED')
        success_rate = (success_count / len(releases) * 100) if releases else 100
        
        return {
            'period': {'start': start_time, 'end': end_time},
            'total_releases': len(releases),
            'total_rollbacks': len(rollbacks),
            'success_rate': round(success_rate, 2),
            'status_distribution': status_stats,
            'risk_distribution': risk_stats,
            'avg_rollback_defects': 
                round(sum(r.get('estimated_defect_count', 0) for r in rollbacks) / len(rollbacks), 0)
                if rollbacks else 0
        }


def get_query_export_manager() -> QueryExportManager:
    """获取查询导出管理器"""
    return QueryExportManager()
